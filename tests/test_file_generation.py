"""Tests for file generation (CSV and XLSX writers)."""

from __future__ import annotations

import csv
from pathlib import Path
from unittest.mock import patch

import pytest
from openpyxl import Workbook, load_workbook

from b2b_data_bridge.config import FileConfig
from b2b_data_bridge.models import ExternalProductRow
from b2b_data_bridge.files import (
    archive_file,
    parse_csv,
    parse_file,
    parse_xlsx,
    quarantine_file,
    write_csv,
    write_file,
    write_xlsx,
)


@pytest.fixture
def file_config() -> FileConfig:
    return FileConfig(encoding="utf-8", csv_delimiter=";", csv_quotechar='"')


@pytest.fixture
def sample_product_rows() -> list[ExternalProductRow]:
    return [
        ExternalProductRow(
            ArticleNumber="SKU-001", ArticleName="Mouse", EAN="4006381333931",
            Description="Wireless mouse", Category="Peripherals", Brand="TechCorp",
            WeightKG="0.12", LastUpdate="2026-03-30T10:00:00",
        ),
        ExternalProductRow(
            ArticleNumber="SKU-002", ArticleName="Hub", EAN="4006381333948",
            Description="USB-C hub", Category="Accessories", Brand="TechCorp",
            WeightKG="0.25", LastUpdate="2026-03-30T10:00:00",
        ),
    ]


class TestCsvWriter:
    def test_write_csv_creates_file(
        self, tmp_path: Path, sample_product_rows: list[ExternalProductRow], file_config: FileConfig,
    ) -> None:
        out = tmp_path / "test.csv"
        result = write_csv(sample_product_rows, out, file_config)
        assert result.exists()
        assert result == out

    def test_csv_content_correct(
        self, tmp_path: Path, sample_product_rows: list[ExternalProductRow], file_config: FileConfig,
    ) -> None:
        out = tmp_path / "test.csv"
        write_csv(sample_product_rows, out, file_config)

        with open(out, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f, delimiter=";")
            rows = list(reader)
        assert len(rows) == 2
        assert rows[0]["ArticleNumber"] == "SKU-001"
        assert rows[1]["EAN"] == "4006381333948"

    def test_csv_utf8_encoding(
        self, tmp_path: Path, file_config: FileConfig,
    ) -> None:
        rows = [
            ExternalProductRow(
                ArticleNumber="SKU-Ü01", ArticleName="Ärmel Mütze", EAN="4006381333931",
            ),
        ]
        out = tmp_path / "utf8.csv"
        write_csv(rows, out, file_config)
        content = out.read_text(encoding="utf-8")
        assert "Ärmel Mütze" in content

    def test_csv_empty_rows_raises(
        self, tmp_path: Path, file_config: FileConfig,
    ) -> None:
        with pytest.raises(ValueError, match="zero rows"):
            write_csv([], tmp_path / "empty.csv", file_config)


class TestXlsxWriter:
    def test_write_xlsx_creates_file(
        self, tmp_path: Path, sample_product_rows: list[ExternalProductRow],
    ) -> None:
        out = tmp_path / "test.xlsx"
        result = write_xlsx(sample_product_rows, out)
        assert result.exists()

    def test_xlsx_content_correct(
        self, tmp_path: Path, sample_product_rows: list[ExternalProductRow],
    ) -> None:
        out = tmp_path / "test.xlsx"
        write_xlsx(sample_product_rows, out)

        wb = load_workbook(str(out))
        ws = wb.active
        assert ws is not None
        rows = list(ws.iter_rows(values_only=True))
        # Header + 2 data rows
        assert len(rows) == 3
        assert rows[0][0] == "ArticleNumber"
        assert rows[1][0] == "SKU-001"
        wb.close()

    def test_xlsx_empty_rows_raises(self, tmp_path: Path) -> None:
        with pytest.raises(ValueError, match="zero rows"):
            write_xlsx([], tmp_path / "empty.xlsx")


class TestWriteFileDispatch:
    def test_write_file_csv(
        self, tmp_path: Path, sample_product_rows: list[ExternalProductRow], file_config: FileConfig,
    ) -> None:
        out = tmp_path / "out.csv"
        result = write_file(sample_product_rows, out, file_config)
        assert result.suffix == ".csv"
        assert result.exists()

    def test_write_file_xlsx(
        self, tmp_path: Path, sample_product_rows: list[ExternalProductRow], file_config: FileConfig,
    ) -> None:
        out = tmp_path / "out.xlsx"
        result = write_file(sample_product_rows, out, file_config)
        assert result.suffix == ".xlsx"
        assert result.exists()


class TestParseCsv:
    def test_parse_csv_with_missing_required_fields(
        self, tmp_path: Path, file_config: FileConfig,
    ) -> None:
        # CSV missing required ArticleName column → pydantic validation error per row
        path = tmp_path / "missing_col.csv"
        path.write_text("ArticleNumber;EAN\nSKU-001;4006381333931\n", encoding="utf-8")

        rows, errors = parse_csv(path, ExternalProductRow, file_config)

        assert len(rows) == 0
        assert len(errors) > 0

    def test_parse_csv_row_limit_truncates(
        self, tmp_path: Path, file_config: FileConfig,
    ) -> None:
        path = tmp_path / "big.csv"
        lines = ["ArticleNumber;ArticleName;EAN"]
        for i in range(5):
            lines.append(f"SKU-{i:03d};Name {i};1234567890123")
        path.write_text("\n".join(lines), encoding="utf-8")

        with patch("b2b_data_bridge.files._MAX_PARSE_ROWS", 2):
            rows, errors = parse_csv(path, ExternalProductRow, file_config)

        assert len(rows) == 2


class TestParseXlsx:
    def _make_xlsx(self, tmp_path: Path, rows: list[list]) -> Path:
        wb = Workbook()
        ws = wb.active
        for row in rows:
            ws.append(row)
        path = tmp_path / "test.xlsx"
        wb.save(str(path))
        wb.close()
        return path

    def test_parse_xlsx_valid_rows(self, tmp_path: Path) -> None:
        path = self._make_xlsx(tmp_path, [
            ["ArticleNumber", "ArticleName", "EAN"],
            ["SKU-001", "Mouse", "4006381333931"],
            ["SKU-002", "Hub", "4006381333948"],
        ])
        rows, errors = parse_xlsx(path, ExternalProductRow)
        assert len(rows) == 2
        assert len(errors) == 0
        assert rows[0].ArticleNumber == "SKU-001"

    def test_parse_xlsx_with_missing_required_fields(self, tmp_path: Path) -> None:
        # Missing ArticleName column → pydantic validation error
        path = self._make_xlsx(tmp_path, [
            ["ArticleNumber", "EAN"],
            ["SKU-001", "4006381333931"],
        ])
        rows, errors = parse_xlsx(path, ExternalProductRow)
        assert len(rows) == 0
        assert len(errors) > 0


class TestParseFileDispatch:
    def test_parse_file_dispatches_xlsx(self, tmp_path: Path, file_config: FileConfig) -> None:
        wb = Workbook()
        ws = wb.active
        ws.append(["ArticleNumber", "ArticleName", "EAN"])
        ws.append(["SKU-001", "Mouse", "4006381333931"])
        xlsx_path = tmp_path / "test.xlsx"
        wb.save(str(xlsx_path))
        wb.close()

        rows, errors = parse_file(xlsx_path, ExternalProductRow, file_config)
        assert len(rows) == 1
        assert rows[0].ArticleNumber == "SKU-001"


class TestArchiveAndQuarantine:
    def test_archive_file_moves_file(self, tmp_path: Path) -> None:
        src = tmp_path / "data.csv"
        src.write_text("content")
        archive_dir = str(tmp_path / "archive")

        result = archive_file(src, archive_dir)

        assert result.exists()
        assert not src.exists()

    def test_archive_collision_creates_unique_name(self, tmp_path: Path) -> None:
        archive_dir = str(tmp_path / "archive")

        src1 = tmp_path / "data.csv"
        src1.write_text("first")
        result1 = archive_file(src1, archive_dir)

        src2 = tmp_path / "data.csv"
        src2.write_text("second")
        result2 = archive_file(src2, archive_dir)

        assert result1 != result2
        assert result1.exists()
        assert result2.exists()

    def test_quarantine_file_moves_file(self, tmp_path: Path) -> None:
        src = tmp_path / "bad.csv"
        src.write_text("garbage")
        failed_dir = str(tmp_path / "failed")

        result = quarantine_file(src, failed_dir, "test reason")

        assert result.exists()
        assert not src.exists()

    def test_quarantine_collision_creates_unique_name(self, tmp_path: Path) -> None:
        failed_dir = str(tmp_path / "failed")

        src1 = tmp_path / "bad.csv"
        src1.write_text("first bad")
        result1 = quarantine_file(src1, failed_dir, "reason")

        src2 = tmp_path / "bad.csv"
        src2.write_text("second bad")
        result2 = quarantine_file(src2, failed_dir, "reason")

        assert result1 != result2
        assert result1.exists()
        assert result2.exists()

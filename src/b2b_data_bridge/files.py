"""File handling — CSV/XLSX read, write, naming conventions, archiving.

Everything file-related lives in one place. No need to scatter across
four modules for a batch integration project.
"""

from __future__ import annotations

import csv
import shutil
from datetime import datetime, timezone
from pathlib import Path
from typing import Sequence, Type, TypeVar

from openpyxl import Workbook, load_workbook
from pydantic import BaseModel
from pydantic import ValidationError as PydanticValidationError

from b2b_data_bridge.config import FileConfig, NamingConfig
from b2b_data_bridge.validation import ValidationError

import logging

logger = logging.getLogger(__name__)

T = TypeVar("T", bound=BaseModel)

_MAX_PARSE_ROWS = 500_000  # safety limit to prevent OOM on malicious files


# ---------------------------------------------------------------------------
# Filename generation
# ---------------------------------------------------------------------------

_JOB_PREFIX = {
    "products": "product_prefix",
    "pricing": "pricing_prefix",
    "stock": "stock_prefix",
    "orders": "order_prefix",
}


def make_filename(job_type: str, ext: str, naming: NamingConfig, ts: datetime | None = None) -> str:
    """Generate filename: PRODUCTS_20260330_143000.csv"""
    ts = ts or datetime.now(timezone.utc)
    prefix = getattr(naming, _JOB_PREFIX[job_type])
    return f"{prefix}_{ts.strftime(naming.timestamp_format)}.{ext}"


# ---------------------------------------------------------------------------
# CSV / XLSX writers
# ---------------------------------------------------------------------------


def _to_dicts(rows: Sequence[BaseModel]) -> list[dict]:
    return [r.model_dump() for r in rows]


def write_csv(rows: Sequence[BaseModel], path: Path, config: FileConfig) -> Path:
    """Write pydantic rows to a semicolon-delimited CSV."""
    if not rows:
        raise ValueError("Cannot write CSV with zero rows")
    dicts = _to_dicts(rows)
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding=config.encoding) as f:
        w = csv.DictWriter(f, fieldnames=dicts[0].keys(),
                           delimiter=config.csv_delimiter, quotechar=config.csv_quotechar,
                           quoting=csv.QUOTE_MINIMAL)
        w.writeheader()
        w.writerows(dicts)
    logger.info("Wrote CSV %s (%d rows)", path.name, len(dicts))
    return path


def write_xlsx(rows: Sequence[BaseModel], path: Path) -> Path:
    """Write pydantic rows to an XLSX file."""
    if not rows:
        raise ValueError("Cannot write XLSX with zero rows")
    dicts = _to_dicts(rows)
    keys = list(dicts[0].keys())
    wb = Workbook()
    ws = wb.active
    ws.append(keys)
    for d in dicts:
        ws.append([d[k] for k in keys])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    logger.info("Wrote XLSX %s (%d rows)", path.name, len(dicts))
    return path


def write_file(rows: Sequence[BaseModel], path: Path, config: FileConfig) -> Path:
    """Write rows in the configured format."""
    if path.suffix == ".xlsx":
        return write_xlsx(rows, path)
    return write_csv(rows, path, config)


# ---------------------------------------------------------------------------
# CSV / XLSX parsers
# ---------------------------------------------------------------------------


class FileParseError(Exception):
    """File could not be parsed at all (encoding / structural issue)."""


def parse_csv(path: Path, model: Type[T], config: FileConfig) -> tuple[list[T], list[ValidationError]]:
    """Parse CSV into typed rows. Returns (rows, parse_errors)."""
    rows: list[T] = []
    errors: list[ValidationError] = []
    try:
        with open(path, "r", encoding=config.encoding) as f:
            for idx, raw in enumerate(csv.DictReader(f, delimiter=config.csv_delimiter,
                                                      quotechar=config.csv_quotechar)):
                if idx >= _MAX_PARSE_ROWS:
                    logger.warning("Row limit (%d) reached for %s — truncating", _MAX_PARSE_ROWS, path.name)
                    break
                try:
                    rows.append(model(**raw))
                except PydanticValidationError as e:
                    for err in e.errors():
                        errors.append(ValidationError(
                            idx, str(err.get("loc", ["?"])[0]),
                            str(err.get("input", "")), err.get("msg", "parse error"),
                        ))
    except UnicodeDecodeError as e:
        raise FileParseError(f"Encoding error in {path}: {e}") from e
    except Exception as e:
        raise FileParseError(f"Cannot parse {path}: {e}") from e
    logger.info("Parsed %s: %d rows, %d errors", path.name, len(rows), len(errors))
    return rows, errors


def parse_xlsx(path: Path, model: Type[T]) -> tuple[list[T], list[ValidationError]]:
    """Parse XLSX into typed rows. Returns (rows, parse_errors)."""
    rows: list[T] = []
    errors: list[ValidationError] = []
    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
        ws = wb.active
        headers: list[str] = []
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            if idx == 0:
                headers = [str(c) if c is not None else "" for c in row]
                continue
            if idx > _MAX_PARSE_ROWS:
                logger.warning("Row limit (%d) reached for %s — truncating", _MAX_PARSE_ROWS, path.name)
                break
            raw = {headers[i]: (str(c) if c is not None else "") for i, c in enumerate(row)}
            try:
                rows.append(model(**raw))
            except PydanticValidationError as e:
                for err in e.errors():
                    errors.append(ValidationError(
                        idx - 1, str(err.get("loc", ["?"])[0]),
                        str(err.get("input", "")), err.get("msg", "parse error"),
                    ))
        wb.close()
    except Exception as e:
        raise FileParseError(f"Cannot parse {path}: {e}") from e
    logger.info("Parsed %s: %d rows, %d errors", path.name, len(rows), len(errors))
    return rows, errors


def parse_file(path: Path, model: Type[T], config: FileConfig) -> tuple[list[T], list[ValidationError]]:
    """Auto-detect format and parse."""
    if path.suffix.lower() == ".xlsx":
        return parse_xlsx(path, model)
    return parse_csv(path, model, config)


# ---------------------------------------------------------------------------
# Archive / quarantine
# ---------------------------------------------------------------------------


def archive_file(file_path: Path, archive_dir: str) -> Path:
    """Move processed file into a date-stamped archive folder."""
    dest_dir = Path(archive_dir) / datetime.now(timezone.utc).strftime("%Y-%m-%d")
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest = dest_dir / file_path.name
    if dest.exists():
        dest = dest_dir / f"{file_path.stem}_{datetime.now(timezone.utc).strftime('%H%M%S%f')}{file_path.suffix}"
    shutil.move(str(file_path), str(dest))
    logger.info("Archived %s → %s", file_path.name, dest)
    return dest


def quarantine_file(file_path: Path, failed_dir: str, reason: str = "") -> Path:
    """Move bad file to quarantine folder."""
    dest_dir = Path(failed_dir) / datetime.now(timezone.utc).strftime("%Y-%m-%d")
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest = dest_dir / file_path.name
    if dest.exists():
        dest = dest_dir / f"{file_path.stem}_{datetime.now(timezone.utc).strftime('%H%M%S%f')}{file_path.suffix}"
    shutil.move(str(file_path), str(dest))
    logger.warning("Quarantined %s (reason: %s)", file_path.name, reason)
    return dest
